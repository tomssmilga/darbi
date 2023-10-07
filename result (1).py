from openpyxl import Workbook, load_workbook
wb=load_workbook('tests/test1.xlsx')
ws=wb.active
max_row=ws.max_row
total=0
for i in range(2,max_row+1):
    id=ws['A' + str(i)].value
    hour=ws['B' + str(i)].value
    rate=ws['C' + str(i)].value
    if (type(hour)!=str and type(rate)!=str):
        salary=hour*rate
        ws['D' +str(i)].value=salary
        if (salary>3000):
            total = total+1
print(total)
wb.save('tests/test1.xlsx')
wb.close()